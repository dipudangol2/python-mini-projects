import json
import datetime, os
from data_scraper import get_weather_data, write_to_xlsx
from openpyxl import Workbook, load_workbook


workbook_path = r"C:\Users\Acer\OneDrive\Documents\code\python-mini-projects\web-scraping\scraper3\weather_data.xlsx"
if os.path.exists(workbook_path):
    wb = load_workbook(workbook_path)
else:
    wb = Workbook()

CURRENT_YEAR = datetime.datetime.now().year
weather_data = {}
MONTHS = {
    "1": "Jan",
    "2": "Feb",
    "3": "Mar",
    "4": "Apr",
    "5": "May",
    "6": "Jun",
    "7": "Jul",
    "8": "Aug",
    "9": "Sep",
    "10": "Oct",
    "11": "Nov",
    "12": "Dec",
}

url = "https://www.accuweather.com/en/np/thankot/241832/daily-weather-forecast/241832"


weather_data = get_weather_data(url, MONTHS, CURRENT_YEAR)
ws = wb.active
ws.title = "Weather data"
header_columns = [
    "Date",
    "Minimum Temperature",
    "Maximum Temperature",
    "Precipitation",
    "Status",
    "More",
]


write_to_xlsx(weather_data, wb, ws, header_columns)
with open("data.json", "w") as jsonfile:
    jsonfile.write(json.dumps(weather_data, indent=4))
    jsonfile.close()
